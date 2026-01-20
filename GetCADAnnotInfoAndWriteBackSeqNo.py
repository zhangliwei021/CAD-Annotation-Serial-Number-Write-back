# -*- coding: utf-8 -*-

"""
ZwCAD 批量读取标注并回写序号 - 增强版（融合美观GUI）
运行前：pip install pywin32 openpyxl
"""

import os
import sys
import shutil
import time
import pythoncom
import openpyxl
import queue
import threading
import ctypes
from openpyxl.utils import get_column_letter
import win32com.client as win32
from win32com.client import constants as cst
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from tkinter.font import Font

# ==========  用户可改区域  ==========
ZWCAD_EXE = r"C:\Program Files\ZWSOFT\ZWCAD 2023\ZWCAD.exe"
WORK_DIR  = r"D:\CAD标识\标识后"
EXCEL_NAME= "数值表.xlsx"
# 序号文字高度（可调整）
TEXT_HEIGHT = 2.5
# 序号偏移量（Y轴向上偏移，避免遮挡原标注）
TEXT_OFFSET_Y = 3.0
# 支持特殊字符的CAD字体（内置字体，无需额外安装）
SUPPORT_FONT = "gbcbig.shx"  # 备选：hztxt.shx、hzfs.shx
# 使用带括号数字而非带圈数字（避免字体兼容性问题）
USE_BRACKET_NUMBERS = True  # True: 使用(1)(2)(3); False: 使用①②③
# ====================================

# ==========  核心业务逻辑（修复版本）  ==========
def ensure_zwcad():
    """若 ZwCAD 未启动则启动，并返回 Application 对象（修复COM启动问题）"""
    pythoncom.CoInitialize()
    cad = None
    try:
        # 先尝试连接已运行的ZwCAD
        cad = win32.GetActiveObject("ZWCAD.Application")
        log_msg("已连接到运行中的ZwCAD")
    except Exception as e:
        log_msg(f"ZwCAD 未启动，正在尝试启动…错误信息：{str(e)}")
        try:
            # 方法1：使用os.startfile启动（更可靠）
            os.startfile(ZWCAD_EXE)
            log_msg("已通过os.startfile启动ZwCAD")
            
            # 延长等待时间，确保CAD完全加载
            time.sleep(15)
            
            # 尝试多次获取CAD对象
            for i in range(8):
                try:
                    cad = win32.GetActiveObject("ZWCAD.Application")
                    log_msg(f"第{i+1}次尝试连接ZwCAD成功")
                    break
                except:
                    log_msg(f"第{i+1}次尝试连接ZwCAD失败，等待2秒...")
                    time.sleep(2)
            else:
                # 方法2：如果GetActiveObject失败，尝试Dispatch
                try:
                    log_msg("尝试使用Dispatch连接ZwCAD...")
                    cad = win32.Dispatch("ZWCAD.Application")
                    log_msg("使用Dispatch连接ZwCAD成功")
                except Exception as e2:
                    raise Exception(f"所有连接方式都失败：{str(e2)}")
                    
        except Exception as e2:
            raise Exception(f"启动ZwCAD失败：{str(e2)}")
    
    if cad:
        cad.Visible = True
        # 刷新视图，确保后续操作正常
        try:
            if cad.ActiveDocument:
                cad.ActiveDocument.Regen(True)
        except:
            pass  # 可能没有活动文档，忽略错误
    else:
        raise Exception("未能获取ZwCAD应用程序对象")
    
    return cad

def clear_and_create_excel():
    """清空工作目录并新建 Excel"""
    if os.path.exists(WORK_DIR):
        for file_name in os.listdir(WORK_DIR):
            file_path = os.path.join(WORK_DIR, file_name)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                except Exception as e:
                    log_msg(f"无法删除文件 {file_path}: {e}")
    else:
        os.makedirs(WORK_DIR)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "说明"
    ws["A1"] = "本文件由脚本自动生成，请勿手动修改"
    excel_full_path = os.path.join(WORK_DIR, EXCEL_NAME)
    wb.save(excel_full_path)
    wb.close()
    log_msg(f"Excel文件已创建：{excel_full_path}")

def number_to_circle(n: int) -> str:
    """1→①  2→② … 20→⑳  大于20用(21)形式，返回兼容CAD的字符串"""
    # 如果使用带括号数字（避免字体兼容性问题）
    if USE_BRACKET_NUMBERS:
        return f"({n})"
    
    # 原逻辑：使用带圈数字
    if 1 <= n <= 20:
        circle_char = chr(0x245F + n)  # ①②…⑳
        try:
            # 测试编码转换，如果失败则使用括号格式
            test_str = circle_char.encode("gbk", errors="replace").decode("gbk")
            if test_str != '?' and test_str != circle_char:
                return test_str
            else:
                return f"({n})"
        except:
            return f"({n})"
    return f"({n})"

def create_special_text_style(doc, style_name="SpecialCharStyle"):
    """创建或获取特殊文本样式（修复版本）"""
    try:
        # 检查样式是否存在
        for style in doc.TextStyles:
            if style.Name.lower() == style_name.lower():
                return style
        
        # 创建新样式
        new_style = doc.TextStyles.Add(style_name)
        new_style.FontFile = SUPPORT_FONT
        new_style.BigFontFile = ""
        new_style.Height = TEXT_HEIGHT
        log_msg(f"  创建了新的文本样式: {style_name}")
        return new_style
    except Exception as e:
        log_msg(f"  创建/获取文本样式失败：{str(e)}，使用默认样式")
        try:
            return doc.TextStyles.Item(0)  # 返回第一个样式（通常是Standard）
        except:
            return None

def convert_to_numeric(text):
    if not text or not isinstance(text, str):
        return text
    
    # 定义「保留原格式」的特殊符号
    special_symbols = {'R', 'Φ', 'ф', 'Ф', '∮', '+', '/', '-', '±', 'X', 'x'}
    
    # 检查是否包含特殊符号或字母
    has_special_or_alpha = any(char.isalpha() or char in special_symbols for char in text)
    if has_special_or_alpha:
        return text  # 包含字母或特殊符号，直接返回原文本

    # 定义纯数值允许的字符
    numeric_chars = {'0','1','2','3','4','5','6','7','8','9','.', ',', '-'}
    # 检查是否所有字符都在允许范围内
    if not all(char in numeric_chars for char in text.replace(" ", "")):
        return text  # 包含不允许的字符，返回原文本

    # 步骤2：纯数值格式文本，进行数值转换和格式化
    conversion_attempts = [
        lambda x: float(x),
        lambda x: float(x.replace(',', '')),
    ]
    
    for attempt in conversion_attempts:
        try:
            result = attempt(text)
            if isinstance(result, float):
                if result.is_integer():
                    return str(int(result))
                else:
                    return f"{result:.2f}"
            else:
                return str(result)
        except (ValueError, TypeError):
            continue
    
    return text

def collect_annotations(dwg_path, cad):
    """提取标注信息（修复版本）"""
    doc = None
    try:
        # 确保cad有活动文档
        if not cad.ActiveDocument:
            log_msg("  CAD没有活动文档，尝试创建新文档")
            doc = cad.Documents.Add()
        else:
            doc = cad.ActiveDocument
        
        # 打开DWG文件
        doc = cad.Documents.Open(dwg_path)
        time.sleep(2)  # 增加等待时间
        
        ents = []
        model_space = doc.ModelSpace
        
        for entity in model_space:
            txt = None
            x = None
            y = None
            try:
                entity_name = entity.EntityName
                if entity_name in ("AcDbDimension", "AcDbRotatedDimension", "AcDbAlignedDimension", "AcDbRadialDimension", "AcDbDiametricDimension"):
                    txt = str(entity.TextOverride) if hasattr(entity, 'TextOverride') and entity.TextOverride else str(entity.Measurement)
                    pt = entity.TextPosition
                    x, y = pt[0], pt[1]
                elif entity_name in ("AcDbText", "AcDbMText"):
                    txt = str(entity.TextString) if hasattr(entity, 'TextString') else str(getattr(entity, 'Text', ''))
                    pt = entity.InsertionPoint
                    x, y = pt[0], pt[1]
            except Exception as e:
                continue

            if txt and txt.strip() and x is not None and y is not None:
                try:
                    x_2dec = round(float(x), 2)
                    y_2dec = round(float(y), 2)
                    ents.append((txt.strip(), x_2dec, y_2dec))
                except (ValueError, TypeError):
                    continue
        
        log_msg(f"  提取到{len(ents)}条有效标注")
        return ents
    except Exception as e:
        raise Exception(f"提取{dwg_path}标注失败：{str(e)}")
    finally:
        if doc:
            try:
                doc.Close(False)
            except:
                pass

def write_to_excel(sheet_name, data):
    """写入Excel（保持原有逻辑）"""
    try:
        excel_full_path = os.path.join(WORK_DIR, EXCEL_NAME)
        wb = openpyxl.load_workbook(excel_full_path)
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx in range(ws.max_row, 1, -1):
                ws.delete_rows(row_idx)
        else:
            ws = wb.create_sheet(sheet_name)
        
        ws["A1"] = "序号"
        ws["B1"] = "标注内容"
        ws["C1"] = "X"
        ws["D1"] = "Y"

        for idx, (txt, x, y) in enumerate(data, 2):
            converted_txt = convert_to_numeric(txt)
            ws.cell(row=idx, column=2, value=converted_txt)
            
            if isinstance(converted_txt, str):
                cleaned_txt = converted_txt.replace('.', '').replace('-', '').replace(',', '').replace(' ', '')
                if cleaned_txt.isdigit():
                    ws.cell(row=idx, column=2).number_format = "0" if '.' not in converted_txt else "0.00"
            
            ws.cell(row=idx, column=3, value=round(float(x), 2)).number_format = "0.00"
            ws.cell(row=idx, column=4, value=round(float(y), 2)).number_format = "0.00"
            ws.cell(row=idx, column=1, value=number_to_circle(idx-1))
        
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        wb.save(excel_full_path)
        wb.close()
        log_msg(f"  Excel工作表「{sheet_name}」已更新")
    except Exception as e:
        raise Exception(f"写入Excel失败：{str(e)}")

def add_labels_back(dwg_path, cad):
    """回写序号到DWG（完全重写修复版本）"""
    doc = None
    try:
        excel_full_path = os.path.join(WORK_DIR, EXCEL_NAME)
        sheet_name = os.path.basename(dwg_path)[:-4]
        
        wb = openpyxl.load_workbook(excel_full_path)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise Exception(f"Excel中不存在工作表「{sheet_name}」")
        
        ws = wb[sheet_name]
        
        # 确保CAD有活动文档
        if not cad.ActiveDocument:
            doc = cad.Documents.Add()
        else:
            doc = cad.ActiveDocument
        
        # 打开DWG文件
        doc = cad.Documents.Open(dwg_path)
        time.sleep(2)
        
        # 创建或使用文本样式
        special_text_style = create_special_text_style(doc)
        style_name = special_text_style.Name if special_text_style else "Standard"
        
        row = 2
        write_count = 0
        
        # 预先创建插入点变量
        insertion_point = None
        
        while True:
            seq_txt = ws.cell(row=row, column=1).value
            x_val = ws.cell(row=row, column=3).value
            y_val = ws.cell(row=row, column=4).value
            
            if seq_txt is None or x_val is None or y_val is None:
                break
            
            seq_txt = str(seq_txt).strip()
            if not seq_txt:
                row += 1
                continue
            
            try:
                x = round(float(x_val), 3)
                y = round(float(y_val), 1)
            except (ValueError, TypeError):
                row += 1
                continue
            
            annotate_y = y + TEXT_OFFSET_Y
            
            # 创建插入点数组
            insertion_point = win32.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, annotate_y, 0.0))
            
            try:
                # 首先尝试使用带括号的数字（确保显示正确）
                bracket_text = f"({row-1})"
                
                # 添加文字
                text_obj = doc.ModelSpace.AddText(bracket_text, insertion_point, TEXT_HEIGHT)
                if text_obj:
                    text_obj.StyleName = style_name
                    text_obj.Color = 1  # 红色
                    text_obj.Update()
                    write_count += 1
                    log_msg(f"  第{row}行序号写入成功: {bracket_text}")
                else:
                    raise Exception("AddText返回None")
                    
            except Exception as e:
                log_msg(f"  第{row}行序号写入失败：{str(e)}")
                # 备选方法：使用简单数字
                try:
                    backup_seq_txt = f"{row-1}"
                    insertion_point = win32.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, annotate_y, 0.0))
                    text_obj = doc.ModelSpace.AddText(backup_seq_txt, insertion_point, TEXT_HEIGHT)
                    if text_obj:
                        text_obj.StyleName = style_name
                        text_obj.Color = 1
                        text_obj.Update()
                        write_count += 1
                        log_msg(f"  第{row}行使用备选序号成功: {backup_seq_txt}")
                except Exception as e2:
                    log_msg(f"  备选方法也失败：{str(e2)}")
            
            row += 1
        
        # 刷新视图
        try:
            doc.Regen(True)
        except:
            pass
        
        # 保存文件
        dwg_filename = os.path.basename(dwg_path)
        new_dwg_path = os.path.join(WORK_DIR, dwg_filename)
        try:
            doc.SaveAs(new_dwg_path)
            log_msg(f"  DWG文件已保存到: {new_dwg_path}")
        except Exception as e:
            raise Exception(f"  文档另存为失败：{str(e)}")
        
        wb.close()
        log_msg(f"  成功回写{write_count}个序号到DWG文件")
        return write_count
        
    except Exception as e:
        raise Exception(f"回写{dwg_path}序号失败：{str(e)}")
    finally:
        if doc:
            try:
                doc.Close(False)
            except:
                pass

def open_output_folder():
    """打开输出文件夹"""
    try:
        if os.path.exists(WORK_DIR):
            os.startfile(WORK_DIR)
            log_msg(f"已打开输出文件夹: {WORK_DIR}")
        else:
            messagebox.showwarning("警告", f"文件夹不存在: {WORK_DIR}")
    except Exception as e:
        messagebox.showerror("错误", f"无法打开文件夹: {str(e)}")

# ==========  全局变量（用于日志队列传递）  ==========
log_queue = None

def log_msg(msg):
    """写入日志到队列（线程安全）"""
    if log_queue and not log_queue.full():
        log_queue.put(("LOG", msg))

# ==========  后台处理线程  ==========
def run_process_async(dwg_files, log_q, status_q):
    """后台执行批量处理任务（不阻塞GUI主线程）"""
    global log_queue
    log_queue = log_q
    cad = None
    success_count = 0
    total_files = len(dwg_files)
    
    try:
        # 初始化通知
        status_q.put(("STATUS", "✅ 开始初始化，清空并创建Excel文件…"))
        status_q.put(("PROGRESS", 5))
        
        # 清空并创建Excel
        clear_and_create_excel()
        time.sleep(0.5)
        
        status_q.put(("STATUS", "🔧 正在连接/启动ZwCAD…"))
        status_q.put(("PROGRESS", 10))
        
        # 启动/连接ZwCAD
        cad = ensure_zwcad()
        time.sleep(1)
        
        # 批量处理DWG文件
        for i, dwg in enumerate(dwg_files):
            current_file_num = i + 1
            progress = 10 + (i / total_files) * 80  # 10%~90% 分配给文件处理
            dwg_name = os.path.basename(dwg)
            
            # 更新进度和状态
            status_q.put(("STATUS", f"📄 正在处理第 {current_file_num}/{total_files} 个文件：{dwg_name}"))
            status_q.put(("PROGRESS", progress))
            log_msg(f"\n===== 开始处理：{dwg_name} =====")
            
            try:
                # 提取标注
                data = collect_annotations(dwg, cad)
                if not data:
                    log_msg("  ⚠️  无有效标注，跳过回写")
                    continue
                
                # 写入Excel
                sheet_name = os.path.basename(dwg)[:-4]
                write_to_excel(sheet_name, data)
                
                # 回写序号
                add_result = add_labels_back(dwg, cad)
                if add_result > 0:
                    success_count += 1
                
            except Exception as e:
                error_msg = f"  ❌ 处理失败：{str(e)}"
                log_msg(error_msg)
                status_q.put(("STATUS", f"❌ 第 {current_file_num} 个文件处理失败：{dwg_name}"))
                continue
        
        # 处理完成
        final_progress = 100
        final_status = f"✅ 批量处理完成！成功 {success_count}/{total_files} 个文件"
        status_q.put(("PROGRESS", final_progress))
        status_q.put(("STATUS", final_status))
        status_q.put(("MESSAGE", ("info", "处理完成",
                                  f"批量处理完成！\n\n成功处理：{success_count}/{total_files} 个文件\n结果保存至：{WORK_DIR}")))
        
        # 自动打开输出文件夹
        open_output_folder()
        
    except Exception as e:
        error_msg = f"❌ 全局任务失败：{str(e)}"
        log_msg(error_msg)
        status_q.put(("STATUS", error_msg))
        status_q.put(("PROGRESS", 0))
        status_q.put(("MESSAGE", ("error", "严重错误", f"程序运行出错：{str(e)}")))
    finally:
        # 关闭ZwCAD
        if cad:
            try:
                cad.Quit()
                log_msg("ZwCAD 已正常关闭")
            except:
                log_msg("⚠️  ZwCAD 关闭失败，需手动关闭")
        # 标记任务完成
        status_q.put(("DONE", None))

# ==========  GUI界面类  ==========
class ZwCADBatchProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("ZwCAD 批量标注提取与序号回写工具")
        self.root.geometry("900x700")
        self.root.configure(bg='#f8f9fa')
        
        # DPI适配（Windows）
        if sys.platform.startswith('win'):
            try:
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass
        
        # 初始化队列
        self.log_queue = queue.Queue(maxsize=1000)
        self.status_queue = queue.Queue(maxsize=100)
        global log_queue
        log_queue = self.log_queue
        
        # 窗口居中
        self.center_window()
        
        # 创建UI
        self.create_ui()
        
        # 启动队列轮询（更新GUI）
        self.check_queues()
        
        # 保存DWG文件列表
        self.dwg_files = []

    def center_window(self):
        """窗口居中显示"""
        self.root.update_idletasks()
        w, h = 900, 700
        x = (self.root.winfo_screenwidth() // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f'{w}x{h}+{x}+{y}')

    def create_ui(self):
        """创建美观的GUI界面"""
        # 1. 标题区域
        title_label = tk.Label(self.root, text="ZwCAD 批量标注提取与序号回写工具",
                               font=("Microsoft YaHei", 22, "bold"),
                               bg='#f8f9fa', fg='#495057')
        title_label.pack(pady=20)
        
        desc_label = tk.Label(self.root, text="批量提取DWG标注信息，生成Excel报表，并回写序号到图纸",
                              font=("Microsoft YaHei", 12),
                              bg='#f8f9fa', fg='#6c757d')
        desc_label.pack(pady=0, padx=20)
        
        # 2. 按钮框架
        btn_frame = tk.Frame(self.root, bg='#f8f9fa')
        btn_frame.pack(pady=30)
        
        # 选择DWG按钮
        self.select_btn = tk.Button(btn_frame, text="选择DWG文件",
                                    font=("Microsoft YaHei", 12),
                                    bg='#0d6efd', fg='white',
                                    activebackground='#0b5ed7',
                                    relief='flat', padx=25, pady=8,
                                    command=self.select_dwg_files)
        self.select_btn.pack(side=tk.LEFT, padx=10)
        
        # 开始处理按钮
        self.process_btn = tk.Button(btn_frame, text="开始批量处理",
                                     font=("Microsoft YaHei", 12),
                                     bg='#28a745', fg='white',
                                     activebackground='#218838',
                                     relief='flat', padx=25, pady=8,
                                     command=self.start_process,
                                     state=tk.DISABLED)
        self.process_btn.pack(side=tk.LEFT, padx=10)
        
        # 打开文件夹按钮
        self.folder_btn = tk.Button(btn_frame, text="打开输出文件夹",
                                    font=("Microsoft YaHei", 12),
                                    bg='#ffc107', fg='black',
                                    activebackground='#ffb300',
                                    relief='flat', padx=25, pady=8,
                                    command=open_output_folder)
        self.folder_btn.pack(side=tk.LEFT, padx=10)
        
        # 退出按钮
        self.quit_btn = tk.Button(btn_frame, text="退出程序",
                                  font=("Microsoft YaHei", 12),
                                  bg='#6c757d', fg='white',
                                  activebackground='#5a6268',
                                  relief='flat', padx=25, pady=8,
                                  command=self.root.destroy)
        self.quit_btn.pack(side=tk.LEFT, padx=10)
        
        # 3. 状态和进度条区域
        status_frame = tk.Frame(self.root, bg='#f8f9fa')
        status_frame.pack(pady=10, fill=tk.X, padx=50)
        
        # 状态标签
        self.status_label = tk.Label(status_frame, text="就绪",
                                     font=("Microsoft YaHei", 11),
                                     bg='#f8f9fa', fg='#28a745',
                                     anchor='w')
        self.status_label.pack(fill=tk.X, pady=5)
        
        # 进度条（美化样式）
        self.progress_bar = ttk.Progressbar(status_frame, length=800, mode='determinate')
        # 自定义进度条样式
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Custom.Horizontal.TProgressbar",
                       background='#28a745',
                       troughcolor='#ecf0f1',
                       bordercolor='#bdc3c7',
                       lightcolor='#28a745',
                       darkcolor='#218838')
        self.progress_bar.configure(style="Custom.Horizontal.TProgressbar")
        self.progress_bar.pack(fill=tk.X, pady=10)
        
        # 4. 日志窗口区域
        log_frame = tk.LabelFrame(self.root, text=" 操作日志 ",
                                  font=("Microsoft YaHei", 12, "bold"),
                                  bg='white', fg='#495057',
                                  padx=10, pady=10)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=50, pady=20)
        
        # 日志文本框
        self.log_text = tk.Text(log_frame, font=("Courier New", 10),
                                bg='#f0f8ff', fg='#2c3e50',
                                wrap=tk.WORD, state=tk.DISABLED)
        self.log_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        
        # 滚动条
        log_scroll = tk.Scrollbar(log_frame, command=self.log_text.yview)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scroll.set)
        
        # 初始化日志
        self.append_log("📌 程序初始化完成，等待选择DWG文件...")

    def select_dwg_files(self):
        """选择DWG文件"""
        try:
            files = filedialog.askopenfilenames(title="请选择需要处理的DWG文件",
                                                filetypes=[("DWG文件", "*.dwg")])
            if files:
                self.dwg_files = list(files)
                file_count = len(self.dwg_files)
                status_msg = f"✅ 已选择 {file_count} 个DWG文件，可点击「开始批量处理」"
                self.status_label.config(text=status_msg, fg=self._get_status_color(status_msg))
                self.process_btn.config(state=tk.NORMAL)
                self.append_log(f"📂 已选择 {file_count} 个DWG文件")
            else:
                self.dwg_files = []
                self.process_btn.config(state=tk.DISABLED)
                self.append_log("⚠️  未选择任何DWG文件")
        except Exception as e:
            error_msg = f"❌ 选择文件失败：{str(e)}"
            self.append_log(error_msg)
            self.status_label.config(text=error_msg, fg=self._get_status_color(error_msg))

    def start_process(self):
        """启动后台处理线程"""
        if not self.dwg_files:
            messagebox.showwarning("警告", "请先选择DWG文件")
            return
        
        # 禁用按钮，防止重复点击
        self.select_btn.config(state=tk.DISABLED)
        self.process_btn.config(state=tk.DISABLED)
        self.folder_btn.config(state=tk.DISABLED)
        
        # 清空日志
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        # 重置进度条
        self.progress_bar['value'] = 0
        
        # 启动后台线程
        threading.Thread(target=run_process_async,
                         args=(self.dwg_files, self.log_queue, self.status_queue),
                         daemon=True).start()

    def append_log(self, msg):
        """追加日志到文本框（线程安全）"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)  # 自动滚动到底部
        self.log_text.config(state=tk.DISABLED)

    def check_queues(self):
        """轮询队列，更新GUI（非阻塞，避免卡顿）"""
        # 处理日志队列
        while not self.log_queue.empty():
            try:
                msg_type, content = self.log_queue.get_nowait()
                if msg_type == "LOG":
                    self.append_log(content)
            except queue.Empty:
                break
        
        # 处理状态/进度队列
        while not self.status_queue.empty():
            try:
                msg_type, data = self.status_queue.get_nowait()
                if msg_type == "STATUS":
                    self.status_label.config(text=data, fg=self._get_status_color(data))
                elif msg_type == "PROGRESS":
                    self.progress_bar['value'] = data
                elif msg_type == "MESSAGE":
                    mtype, title, msg = data
                    getattr(messagebox, f"show{mtype}")(title, msg, parent=self.root)
                elif msg_type == "DONE":
                    # 恢复按钮状态
                    self.select_btn.config(state=tk.NORMAL)
                    self.process_btn.config(state=tk.NORMAL)
                    self.folder_btn.config(state=tk.NORMAL)
            except queue.Empty:
                break
        
        # 定时轮询（100ms一次，不阻塞GUI）
        self.root.after(100, self.check_queues)

    def _get_status_color(self, text):
        """状态文本颜色区分"""
        if text.startswith("✅"):
            return "#28a745"  # 成功-绿色
        elif text.startswith("❌"):
            return "#dc3545"  # 错误-红色
        elif text.startswith("⚠️"):
            return "#ffc107"  # 警告-黄色
        elif text.startswith("🔧") or text.startswith("📄"):
            return "#0d6efd"  # 处理中-蓝色
        else:
            return "#495057"  # 普通-深灰色

# ==========  程序入口  ==========
if __name__ == '__main__':
    root = tk.Tk()
    app = ZwCADBatchProcessor(root)
    root.mainloop()
